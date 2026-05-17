"""
backtest_runner.py  — V3
═══════════════════════════════════════════════════════════
BTC-only backtest: BTCUSDT 15m (Binance, ~3000 bars)
                   BTC-USD  4h  (yfinance, ~4380 bars)

Change 5 features:
  - Binance paginated fetch for 15m BTC
  - yfinance fetch for 4h BTC (730-day period)
  - Circuit breaker: 3 consecutive losses -> skip 5 bars
  - cvd_move_ratio + pavp_override columns in trade log
  - exhaustion_blocks counter
  - Calibration report with pavp_override / move_ratio<0.5 / exhaustion stats
  - Excel output with per-grade, per-regime, per-hour breakdowns
  - Filters: skip grade C, skip BREAKOUT regime, skip opposing momentum
  - Min confidence: A>=80, B>=65

RR ratio: 1.8
"""

import sys
import os
import json
import math
import time
import warnings
import requests
import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# ── add project root to path ──────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from indicator_engine  import compute_all_indicators
from signal_enrichment import enrich_signal
from scoring_engine_py import score_signal

# ─────────────────────────────────────────────────────────
# FILTER CONSTANTS
# ─────────────────────────────────────────────────────────
FILTER_SKIP_GRADE_C       = True
FILTER_SKIP_BREAKOUT      = True
FILTER_SKIP_OPPOSING_MOM  = True
FILTER_MIN_CONF           = {"A": 80.0, "B": 65.0, "C": 0.0, "NONE": 0.0}

RR_RATIO                  = 1.8
WINDOW                    = 100        # bars of history per signal
STEP                      = 10         # evaluate every N bars (speed vs granularity)
CIRCUIT_BREAKER_LOSSES    = 3          # consecutive losses before skip
CIRCUIT_BREAKER_SKIP      = 5          # bars to skip after trigger


# ─────────────────────────────────────────────────────────
# DATA FETCHERS
# ─────────────────────────────────────────────────────────

def _fetch_binance_paginated(symbol: str, interval: str, total_bars: int) -> pd.DataFrame:
    """Fetch up to total_bars from Binance public klines API (max 1000/request)."""
    url     = "https://api.binance.com/api/v3/klines"
    bars    = []
    end_ms  = None
    limit   = 1000

    print(f"  Fetching {total_bars} bars from Binance ({symbol} {interval})...")

    while len(bars) < total_bars:
        params = {"symbol": symbol, "interval": interval, "limit": min(limit, total_bars - len(bars))}
        if end_ms:
            params["endTime"] = end_ms

        try:
            r = requests.get(url, params=params, timeout=15)
            r.raise_for_status()
            chunk = r.json()
        except Exception as e:
            print(f"  Binance fetch error: {e}")
            break

        if not chunk:
            break

        bars = chunk + bars          # prepend (oldest first)
        end_ms = int(chunk[0][0]) - 1
        time.sleep(0.2)

        if len(chunk) < limit:
            break

    if not bars:
        raise RuntimeError("Binance returned no data")

    df = pd.DataFrame(bars, columns=[
        "open_time","open","high","low","close","volume",
        "close_time","qav","num_trades","tbbav","tbqav","ignore"
    ])
    df["time"]   = pd.to_datetime(df["open_time"].astype(float), unit="ms")
    df["open"]   = df["open"].astype(float)
    df["high"]   = df["high"].astype(float)
    df["low"]    = df["low"].astype(float)
    df["close"]  = df["close"].astype(float)
    df["volume"] = df["volume"].astype(float)

    df = df[["time","open","high","low","close","volume"]].sort_values("time").reset_index(drop=True)
    print(f"  Got {len(df)} bars  ({df['time'].iloc[0].date()} to {df['time'].iloc[-1].date()})")
    return df


def _fetch_yahoo(symbol: str, interval: str, period: str) -> pd.DataFrame:
    """Fetch via yfinance with SSL workaround."""
    try:
        import yfinance as yf
        try:
            from curl_cffi.requests import Session
            session = Session(verify=False, impersonate="chrome110")
            ticker  = yf.Ticker(symbol, session=session)
        except ImportError:
            ticker = yf.Ticker(symbol)

        df = ticker.history(period=period, interval=interval, auto_adjust=True)
        df = df.reset_index()

        time_col = "Datetime" if "Datetime" in df.columns else "Date"
        df = df.rename(columns={
            time_col: "time",
            "Open":   "open",
            "High":   "high",
            "Low":    "low",
            "Close":  "close",
            "Volume": "volume",
        })
        df["time"] = pd.to_datetime(df["time"]).dt.tz_localize(None)
        df = df[["time","open","high","low","close","volume"]].dropna()
        df = df.sort_values("time").reset_index(drop=True)
        print(f"  Got {len(df)} bars  ({df['time'].iloc[0].date()} to {df['time'].iloc[-1].date()})")
        return df
    except Exception as e:
        raise RuntimeError(f"yfinance fetch failed: {e}")


def fetch_ohlcv(symbol: str, timeframe: str, source: str = "auto") -> pd.DataFrame:
    """
    Unified fetch entry point.
    source='binance' -> Binance REST
    source='yahoo'   -> yfinance
    source='auto'    -> binance for 15m, yahoo for 4h
    """
    print(f"\nFetching {symbol} {timeframe} ...")
    if source == "auto":
        source = "binance" if timeframe in ("1m","5m","15m","30m") else "yahoo"

    if source == "binance":
        sym = symbol.replace("-","").upper()
        if not sym.endswith("USDT"):
            sym = sym.replace("USD","USDT")
        return _fetch_binance_paginated(sym, timeframe, total_bars=3000)
    else:
        yf_sym = symbol if "-" in symbol else symbol.replace("USDT","-USD")
        period = "730d" if timeframe in ("1h","2h","4h","1d") else "60d"
        return _fetch_yahoo(yf_sym, timeframe, period)


# ─────────────────────────────────────────────────────────
# BACKTEST CORE
# ─────────────────────────────────────────────────────────

def run_backtest(
    df:         pd.DataFrame,
    symbol:     str,
    timeframe:  str,
    rr:         float = RR_RATIO,
    window:     int   = WINDOW,
    step:       int   = STEP,
) -> tuple:
    """
    Bar-by-bar backtest.

    Returns:
        (trade_log, total_signals, no_trade_count,
         filter_count, exhaustion_blocks)
    """
    closes    = df["close"].values
    highs     = df["high"].values
    lows      = df["low"].values
    times     = df["time"].values
    n         = len(df)

    trade_log        = []
    total_signals    = 0
    no_trade_count   = 0
    filter_count     = 0
    exhaustion_blocks = 0

    consecutive_losses = 0
    skip_until         = -1

    for i in range(window, n - 1, step):

        # Circuit breaker
        if i < skip_until:
            continue

        slice_df = df.iloc[i - window: i].copy().reset_index(drop=True)
        ts       = pd.Timestamp(times[i])
        hour_of_day = ts.hour

        try:
            snap     = compute_all_indicators(slice_df)
            enriched = enrich_signal(
                indicator_snapshot = snap,
                symbol    = symbol,
                timeframe = timeframe,
                price     = {"open":  float(slice_df["open"].iloc[-1]),
                             "high":  float(slice_df["high"].iloc[-1]),
                             "low":   float(slice_df["low"].iloc[-1]),
                             "close": float(closes[i - 1])},
                volume    = float(slice_df["volume"].iloc[-1]),
            )
            decision = score_signal(enriched)
        except Exception as e:
            continue

        total_signals += 1

        # Track exhaustion blocks
        if decision.get("blocked_by_exhaustion"):
            exhaustion_blocks += 1
            no_trade_count    += 1
            continue

        bias  = decision.get("bias",        "NO TRADE")
        grade = decision.get("setup_grade", "NONE")
        conf  = decision.get("confidence",  0.0)
        regime = decision.get("regime",     "UNCERTAIN")
        sb    = decision.get("score_breakdown", {})
        pavp_override = decision.get("pavp_override", False)

        snap_cvd   = snap.get("cvd", {})
        cvd_mr     = snap_cvd.get("delta_implied", {}).get("move_ratio", 1.0)

        # ── Filters ──────────────────────────────────────
        if bias == "NO TRADE":
            no_trade_count += 1
            continue

        if FILTER_SKIP_GRADE_C and grade in ("C","NONE"):
            filter_count += 1
            continue

        if conf < FILTER_MIN_CONF.get(grade, 0.0):
            filter_count += 1
            continue

        if FILTER_SKIP_BREAKOUT and regime == "BREAKOUT":
            filter_count += 1
            continue

        direction = "LONG" if "LONG" in bias else "SHORT"

        if FILTER_SKIP_OPPOSING_MOM:
            ts_dir    = snap.get("trend_speed", {}).get("direction", "FLAT")
            zz_struct = snap.get("zigzag",      {}).get("structure", "NEUTRAL")
            mom_conflict = (
                (direction == "LONG"  and ts_dir == "BEARISH" and zz_struct == "BEARISH") or
                (direction == "SHORT" and ts_dir == "BULLISH" and zz_struct == "BULLISH")
            )
            if mom_conflict:
                filter_count += 1
                continue

        # ── Trade simulation ──────────────────────────────
        entry_price = float(closes[i])
        atr_val     = snap.get("atr", {}).get("atr_value", entry_price * 0.002)

        if direction == "LONG":
            stop   = entry_price - atr_val * 1.5
            target = entry_price + atr_val * 1.5 * rr
        else:
            stop   = entry_price + atr_val * 1.5
            target = entry_price - atr_val * 1.5 * rr

        outcome  = "TIMEOUT"
        bars_to  = 0
        max_bars = 50

        for j in range(i + 1, min(i + 1 + max_bars, n)):
            h = float(highs[j])
            l = float(lows[j])
            bars_to = j - i

            if direction == "LONG":
                if l <= stop:   outcome = "LOSS"; break
                if h >= target: outcome = "WIN";  break
            else:
                if h >= stop:   outcome = "LOSS"; break
                if l <= target: outcome = "WIN";  break

        # Circuit breaker update
        if outcome == "LOSS":
            consecutive_losses += 1
            if consecutive_losses >= CIRCUIT_BREAKER_LOSSES:
                skip_until         = i + CIRCUIT_BREAKER_SKIP
                consecutive_losses = 0
        elif outcome == "WIN":
            consecutive_losses = 0

        trade_log.append({
            "bar_index":        i,
            "timestamp":        str(ts),
            "hour_of_day":      hour_of_day,
            "symbol":           symbol,
            "timeframe":        timeframe,
            "direction":        direction,
            "grade":            grade,
            "confidence":       round(float(conf), 1),
            "regime":           regime,
            "score_structure":  sb.get("structure",  0.0),
            "score_location":   sb.get("location",   0.0),
            "score_momentum":   sb.get("momentum",   0.0),
            "score_orderflow":  sb.get("order_flow", 0.0),
            "score_total":      sb.get("total",      0.0),
            "cvd_move_ratio":   round(float(cvd_mr), 4),
            "pavp_override":    bool(pavp_override),
            "entry_price":      round(float(entry_price), 6),
            "target_price":     round(float(target),      6),
            "stop_price":       round(float(stop),        6),
            "outcome":          outcome,
            "bars_to_outcome":  bars_to,
        })

    return trade_log, total_signals, no_trade_count, filter_count, exhaustion_blocks


# ─────────────────────────────────────────────────────────
# CALIBRATION REPORT
# ─────────────────────────────────────────────────────────

def build_calibration_report(
    trade_log:        list,
    symbol:           str,
    timeframe:        str,
    total_bars:       int,
    total_signals:    int,
    no_trade_count:   int,
    filter_count:     int,
    exhaustion_blocks:int,
    rr:               float,
) -> dict:

    df = pd.DataFrame(trade_log)
    total_trades = len(df)

    def grade_stats(g):
        sub = df[df.grade == g]
        wins   = (sub.outcome == "WIN").sum()
        losses = (sub.outcome == "LOSS").sum()
        n      = len(sub)
        wr     = round(wins / n * 100, 1) if n else 0.0
        avg_c  = round(sub.confidence.mean(), 1) if n else 0.0
        exp    = round(wr/100 * rr - (1 - wr/100), 3) if n else -1.0
        return {"count":int(n),"wins":int(wins),"losses":int(losses),
                "win_rate":wr,"avg_confidence":avg_c,"expectancy_r":exp}

    def regime_stats(r):
        sub = df[df.regime == r]
        n   = len(sub)
        wr  = round((sub.outcome=="WIN").sum() / n * 100, 1) if n else 0.0
        return {"count":int(n),"win_rate":wr}

    def dir_stats(d):
        sub = df[df.direction == d]
        n   = len(sub)
        wr  = round((sub.outcome=="WIN").sum() / n * 100, 1) if n else 0.0
        return {"count":int(n),"win_rate":wr}

    def hour_stats(h):
        sub = df[df.hour_of_day == h]
        n   = len(sub)
        wr  = round((sub.outcome=="WIN").sum() / n * 100, 1) if n else 0.0
        return {"count":int(n),"win_rate":wr}

    def bucket_stats(lo, hi):
        sub = df[(df.score_total >= lo) & (df.score_total < hi)]
        n   = len(sub)
        wr  = round((sub.outcome=="WIN").sum() / n * 100, 1) if n else 0.0
        return {"count":int(n),"win_rate":wr}

    # PAVP override stats
    po = df[df.pavp_override == True]
    po_wins   = (po.outcome == "WIN").sum()
    po_losses = (po.outcome == "LOSS").sum()
    po_n      = len(po)
    po_wr     = round(po_wins / po_n * 100, 1) if po_n else 0.0
    po_conf   = round(po.confidence.mean(), 1) if po_n else 0.0
    po_exp    = round(po_wr/100 * rr - (1 - po_wr/100), 3) if po_n else -1.0

    # CVD move ratio < 0.5 stats
    mr = df[df.cvd_move_ratio < 0.5]
    mr_wins   = (mr.outcome == "WIN").sum()
    mr_losses = (mr.outcome == "LOSS").sum()
    mr_n      = len(mr)
    mr_wr     = round(mr_wins / mr_n * 100, 1) if mr_n else 0.0
    mr_conf   = round(mr.confidence.mean(), 1) if mr_n else 0.0
    mr_exp    = round(mr_wr/100 * rr - (1 - mr_wr/100), 3) if mr_n else -1.0

    return {
        "symbol":        symbol,
        "timeframe":     timeframe,
        "total_bars":    total_bars,
        "total_signals": total_signals,
        "total_trades":  total_trades,
        "rr_ratio":      rr,
        "by_grade": {g: grade_stats(g) for g in ("A","B","C","NONE")},
        "by_regime": {r: regime_stats(r)
                      for r in df.regime.unique() if not pd.isna(r)},
        "by_direction": {d: dir_stats(d) for d in df.direction.unique()},
        "by_hour": {str(h): hour_stats(h) for h in sorted(df.hour_of_day.unique())},
        "score_buckets": {
            "75-100": bucket_stats(75, 101),
            "60-74":  bucket_stats(60,  75),
            "50-59":  bucket_stats(50,  60),
        },
        "pavp_override_stats": {
            "count":int(po_n),"wins":int(po_wins),"losses":int(po_losses),
            "win_rate":po_wr,"avg_confidence":po_conf,"expectancy_r":po_exp,
        },
        "cvd_move_ratio_lt05": {
            "count":int(mr_n),"wins":int(mr_wins),"losses":int(mr_losses),
            "win_rate":mr_wr,"avg_confidence":mr_conf,"expectancy_r":mr_exp,
        },
        "exhaustion_blocks": int(exhaustion_blocks),
        "no_trade_filter_effectiveness": {
            "engine_no_trade":  int(no_trade_count),
            "filter_rejected":  int(filter_count),
            "exhaustion_blocks":int(exhaustion_blocks),
            "total_blocked":    int(no_trade_count + filter_count + exhaustion_blocks),
            "pct_of_total_bars": round(
                (no_trade_count + filter_count + exhaustion_blocks) / max(total_signals, 1) * 100, 1
            ),
        },
        "active_filters": {
            "skip_grade_c":      FILTER_SKIP_GRADE_C,
            "skip_breakout":     FILTER_SKIP_BREAKOUT,
            "skip_opposing_mom": FILTER_SKIP_OPPOSING_MOM,
            "min_conf_A":        FILTER_MIN_CONF["A"],
            "min_conf_B":        FILTER_MIN_CONF["B"],
        },
        "calibration_table": {
            g: grade_stats(g)["win_rate"] for g in ("A","B","C","NONE")
        },
    }


# ─────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────

def write_excel(trade_log: list, report: dict, filepath: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl not installed — skipping Excel output")
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Trade Log ───────────────────────────────
    ws = wb.active
    ws.title = "Trade Log"
    df = pd.DataFrame(trade_log)
    if not df.empty:
        headers = list(df.columns)
        ws.append(headers)
        hdr_fill = PatternFill("solid", fgColor="1E2235")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = Font(color="E8EAF2", bold=True)

        win_fill  = PatternFill("solid", fgColor="0D3B26")
        loss_fill = PatternFill("solid", fgColor="3B0D0D")

        for _, row in df.iterrows():
            ws.append(list(row))
            last = ws.max_row
            outcome = str(row.get("outcome",""))
            if outcome == "WIN":
                for c in ws[last]: c.fill = win_fill
            elif outcome == "LOSS":
                for c in ws[last]: c.fill = loss_fill

        for i, col in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(str(col))+2, 12)

    # ── Sheet 2: Summary ─────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    rows = [
        ["Symbol",      report["symbol"]],
        ["Timeframe",   report["timeframe"]],
        ["Total Bars",  report["total_bars"]],
        ["Total Signals",report["total_signals"]],
        ["Total Trades", report["total_trades"]],
        ["RR Ratio",    report["rr_ratio"]],
        [""],
        ["By Grade"],
    ]
    for g, s in report["by_grade"].items():
        rows.append([f"  Grade {g}", f"Win rate: {s['win_rate']}%",
                     f"Count: {s['count']}", f"Expectancy: {s['expectancy_r']}R"])
    rows.append([""])
    rows.append(["By Direction"])
    for d, s in report["by_direction"].items():
        rows.append([f"  {d}", f"Win rate: {s['win_rate']}%", f"Count: {s['count']}"])
    rows.append([""])
    rows.append(["Exhaustion Blocks", report["exhaustion_blocks"]])
    rows.append(["PAVP Override Win Rate",
                 f"{report['pavp_override_stats']['win_rate']}%",
                 f"n={report['pavp_override_stats']['count']}"])
    rows.append(["CVD Move Ratio <0.5 Win Rate",
                 f"{report['cvd_move_ratio_lt05']['win_rate']}%",
                 f"n={report['cvd_move_ratio_lt05']['count']}"])

    for r in rows:
        ws2.append(r)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 20

    wb.save(filepath)
    print(f"  Excel saved: {filepath}")


# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────

def run_btc_backtest():
    configs = [
        {"symbol": "BTCUSDT", "yf_symbol": "BTC-USD", "timeframe": "15m", "source": "binance"},
    ]

    for cfg in configs:
        tf  = cfg["timeframe"]
        sym = cfg["symbol"]
        src = cfg["source"]

        print(f"\n{'='*60}")
        print(f"  BTC {tf.upper()} BACKTEST")
        print(f"{'='*60}")

        try:
            fetch_sym = cfg["yf_symbol"] if src == "yahoo" else sym
            df = fetch_ohlcv(fetch_sym, tf, source=src)
        except Exception as e:
            print(f"  Data fetch failed: {e}")
            continue

        if len(df) < WINDOW + 10:
            print(f"  Insufficient data ({len(df)} bars). Skipping.")
            continue

        print(f"  Running backtest on {len(df)} bars (window={WINDOW}, step={STEP}, RR={RR_RATIO})...")
        t0 = time.time()

        trade_log, total_sig, no_trade, filtered, exh_blocks = run_backtest(
            df, sym, tf, rr=RR_RATIO, window=WINDOW, step=STEP
        )

        elapsed = time.time() - t0
        n_trades = len(trade_log)
        wins     = sum(1 for t in trade_log if t["outcome"] == "WIN")
        wr       = round(wins / n_trades * 100, 1) if n_trades else 0.0

        print(f"  Done in {elapsed:.1f}s")
        print(f"  Signals evaluated : {total_sig}")
        print(f"  Trades taken      : {n_trades}")
        print(f"  Win rate          : {wr}%  ({wins}W / {n_trades - wins}L)")
        print(f"  Exhaustion blocks : {exh_blocks}")
        print(f"  Filtered out      : {filtered}")

        report = build_calibration_report(
            trade_log, sym, tf,
            total_bars        = len(df),
            total_signals     = total_sig,
            no_trade_count    = no_trade,
            filter_count      = filtered,
            exhaustion_blocks = exh_blocks,
            rr                = RR_RATIO,
        )

        prefix = f"BTCUSDT_{tf}"

        # Save trade log CSV
        csv_path = f"{prefix}_trade_log.csv"
        pd.DataFrame(trade_log).to_csv(csv_path, index=False)
        print(f"  Trade log: {csv_path}")

        # Save calibration report JSON
        json_path = f"{prefix}_calibration_report.json"
        with open(json_path, "w") as f:
            json.dump(report, f, indent=2)
        print(f"  Calibration: {json_path}")

        # Save Excel
        xlsx_path = f"{prefix}_backtest_results.xlsx"
        write_excel(trade_log, report, xlsx_path)

        # Print grade breakdown
        print(f"\n  Grade breakdown:")
        for g, s in report["by_grade"].items():
            if s["count"] > 0:
                print(f"    {g}: {s['win_rate']}% WR  (n={s['count']}, exp={s['expectancy_r']}R)")

        print(f"\n  PAVP override: {report['pavp_override_stats']['win_rate']}% WR"
              f"  (n={report['pavp_override_stats']['count']})")
        print(f"  CVD MR <0.5:  {report['cvd_move_ratio_lt05']['win_rate']}% WR"
              f"  (n={report['cvd_move_ratio_lt05']['count']})")
        print(f"  Breakeven WR needed at {RR_RATIO}R: {100/(1+RR_RATIO):.1f}%")


if __name__ == "__main__":
    run_btc_backtest()

"""
trade_memory.py
═══════════════════════════════════════════════════════════
Persistent trade log with Excel export.

Storage : trades_memory.json (same directory as this file)
Export  : openpyxl-generated .xlsx (Sheet 1 = Trade Log,
          Sheet 2 = Performance Stats)

No database required. JSON file survives server restarts
and is automatically loaded when the server starts.
═══════════════════════════════════════════════════════════
"""

import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

_HERE = os.path.dirname(os.path.abspath(__file__))
MEMORY_FILE = os.path.join(_HERE, "trades_memory.json")


# ─────────────────────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────

def _load() -> List[Dict]:
    if os.path.exists(MEMORY_FILE):
        with open(MEMORY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def _save(trades: List[Dict]) -> None:
    with open(MEMORY_FILE, "w", encoding="utf-8") as f:
        json.dump(trades, f, indent=2, default=str)


# ─────────────────────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────────────────────

def record_trade(trade: Dict[str, Any]) -> Dict[str, Any]:
    """
    Append a trade to the persistent log.
    Adds auto-assigned id and recorded_at timestamp.
    Returns the stored trade dict.
    """
    trades = _load()
    trade = dict(trade)                              # don't mutate caller's dict
    trade["id"]          = len(trades) + 1
    trade["recorded_at"] = datetime.utcnow().isoformat() + "Z"
    trades.append(trade)
    _save(trades)
    return trade


def get_trades(
    symbol: Optional[str] = None,
    limit:  int           = 500,
) -> List[Dict]:
    """Return most-recent `limit` trades, optionally filtered by symbol."""
    trades = _load()
    if symbol:
        trades = [t for t in trades if
                  t.get("symbol", "").upper() == symbol.upper()]
    return trades[-limit:]


def get_stats() -> Dict[str, Any]:
    """
    Aggregate win-rate statistics.
    Trades with outcome PENDING are excluded from rate calculations.
    """
    trades    = _load()
    completed = [t for t in trades if t.get("outcome") in ("WIN", "LOSS", "BREAKEVEN")]
    wins      = [t for t in completed if t.get("outcome") == "WIN"]

    by_asset:  Dict[str, Dict] = {}
    by_signal: Dict[str, Dict] = {}

    for t in completed:
        sym = t.get("symbol",      "UNKNOWN")
        sig = t.get("signal_type", t.get("direction", "UNKNOWN"))
        won = (t.get("outcome") == "WIN")

        for key, bucket in [(sym, by_asset), (sig, by_signal)]:
            if key not in bucket:
                bucket[key] = {"total": 0, "wins": 0}
            bucket[key]["total"] += 1
            if won:
                bucket[key]["wins"] += 1

    for bucket in (by_asset, by_signal):
        for d in bucket.values():
            d["win_rate"] = (
                round(d["wins"] / d["total"] * 100, 1) if d["total"] else 0.0
            )

    n_completed = len(completed)
    return {
        "total":     len(trades),
        "completed": n_completed,
        "wins":      len(wins),
        "win_rate":  round(len(wins) / n_completed * 100, 1) if n_completed else 0.0,
        "by_asset":  by_asset,
        "by_signal": by_signal,
    }


# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────

def export_excel(path: str = "trade_memory_export.xlsx") -> str:
    """
    Generate a styled .xlsx file from the trade log.
    Returns the path where the file was written.
    Raises RuntimeError if openpyxl is not installed.
    """
    if not EXCEL_AVAILABLE:
        raise RuntimeError(
            "openpyxl not installed. Run: pip install openpyxl"
        )

    trades = _load()
    stats  = get_stats()
    wb     = openpyxl.Workbook()

    # ── colours ───────────────────────────────────────────────
    C_BG       = "080C10"
    C_SURFACE  = "0D1219"
    C_SURFACE2 = "141C26"
    C_CYAN     = "00D4FF"
    C_GREEN    = "00D88A"
    C_RED      = "FF4D6D"
    C_AMBER    = "F5A623"
    C_TEXT     = "E8EDF2"
    C_DIM      = "4A6480"

    HDR_FILL   = PatternFill("solid", fgColor=C_SURFACE)
    HDR_FONT   = Font(color=C_CYAN,  bold=True, name="Calibri", size=10)
    HDR_BORDER = Border(
        bottom=Side(style="medium", color=C_CYAN),
        right =Side(style="thin",   color=C_SURFACE2),
    )
    WIN_FILL  = PatternFill("solid", fgColor="071A10")
    LOSS_FILL = PatternFill("solid", fgColor="1A0708")
    PEND_FILL = PatternFill("solid", fgColor=C_SURFACE)
    BE_FILL   = PatternFill("solid", fgColor=C_SURFACE2)

    # ── Sheet 1: Trade Log ────────────────────────────────────
    ws = wb.active
    ws.title = "Trade Log"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C_CYAN

    COLS = [
        ("ID",           "id"),
        ("Recorded",     "recorded_at"),
        ("Symbol",       "symbol"),
        ("Timeframe",    "timeframe"),
        ("Direction",    "direction"),
        ("Signal Type",  "signal_type"),
        ("Grade",        "grade"),
        ("Regime",       "regime"),
        ("Confidence",   "confidence"),
        ("Risk",         "risk"),
        ("Entry Price",  "entry_price"),
        ("Exit Price",   "exit_price"),
        ("Outcome",      "outcome"),
        ("PnL %",        "pnl_pct"),
        ("R Multiple",   "r_multiple"),
        ("Notes",        "notes"),
    ]

    # Header row
    for ci, (label, _) in enumerate(COLS, 1):
        cell            = ws.cell(row=1, column=ci, value=label)
        cell.font       = HDR_FONT
        cell.fill       = HDR_FILL
        cell.border     = HDR_BORDER
        cell.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Data rows
    for ri, trade in enumerate(trades, 2):
        outcome  = trade.get("outcome", "PENDING")
        row_fill = (WIN_FILL  if outcome == "WIN"       else
                    LOSS_FILL if outcome == "LOSS"      else
                    BE_FILL   if outcome == "BREAKEVEN" else PEND_FILL)

        for ci, (_, key) in enumerate(COLS, 1):
            val  = trade.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(name="Calibri", size=9, color=C_TEXT)

            if key == "outcome":
                color = (C_GREEN if outcome == "WIN" else
                         C_RED   if outcome == "LOSS" else
                         C_AMBER if outcome == "PENDING" else C_DIM)
                cell.font = Font(name="Calibri", size=9, bold=True, color=color)

            elif key in ("pnl_pct", "r_multiple") and isinstance(val, (int, float)):
                cell.number_format = "0.00"

        ws.row_dimensions[ri].height = 15

    # Column widths + freeze
    col_widths = {
        "id": 6, "recorded_at": 22, "symbol": 10, "timeframe": 10,
        "direction": 10, "signal_type": 14, "grade": 8, "regime": 12,
        "confidence": 11, "risk": 9, "entry_price": 12, "exit_price": 12,
        "outcome": 11, "pnl_pct": 9, "r_multiple": 11, "notes": 30,
    }
    for ci, (_, key) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(key, 14)

    ws.freeze_panes = "A2"

    # ── Sheet 2: Performance Stats ────────────────────────────
    ws2 = wb.create_sheet("Performance Stats")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = C_GREEN

    ws2["A1"] = "TRADING PERFORMANCE SUMMARY"
    ws2["A1"].font  = Font(bold=True, size=14, color=C_CYAN, name="Calibri")
    ws2["A1"].fill  = PatternFill("solid", fgColor=C_SURFACE)
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 28

    generated = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    ws2["A2"] = generated
    ws2["A2"].font = Font(size=9, color=C_DIM, name="Calibri", italic=True)
    ws2.merge_cells("A2:D2")

    stat_rows: List[tuple] = [
        ("", ""),
        ("OVERVIEW",             ""),
        ("Total Signals Logged", stats["total"]),
        ("Completed Trades",     stats["completed"]),
        ("Wins",                 stats["wins"]),
        ("Win Rate",             f"{stats['win_rate']}%"),
        ("", ""),
        ("WIN RATE BY ASSET",    ""),
    ]
    for sym, data in stats.get("by_asset", {}).items():
        stat_rows.append(
            (f"  {sym}", f"{data['win_rate']}%  ({data['wins']}/{data['total']})")
        )
    stat_rows += [
        ("", ""),
        ("WIN RATE BY SIGNAL TYPE", ""),
    ]
    for sig, data in stats.get("by_signal", {}).items():
        stat_rows.append(
            (f"  {sig}", f"{data['win_rate']}%  ({data['wins']}/{data['total']})")
        )

    for ri, (k, v) in enumerate(stat_rows, 4):
        kcell = ws2.cell(row=ri, column=1, value=k)
        vcell = ws2.cell(row=ri, column=2, value=v)
        is_hdr = bool(k and k == k.upper() and k.strip())
        kcell.font = Font(
            bold=is_hdr, name="Calibri", size=10,
            color=C_CYAN if is_hdr else C_TEXT,
        )
        vcell.font = Font(name="Calibri", size=10, color=C_TEXT)
        if is_hdr:
            kcell.fill = PatternFill("solid", fgColor=C_SURFACE2)
            vcell.fill = PatternFill("solid", fgColor=C_SURFACE2)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 28

    wb.save(path)
    return path
